import asyncio
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, StateFilter
from aiogram.types import (
    InlineKeyboardMarkup, 
    InlineKeyboardButton, 
    WebAppInfo, 
    CallbackQuery,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    FSInputFile,
    BotCommand,
    BotCommandScopeChat
)
from aiogram.types.web_app_data import WebAppData
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from dotenv import load_dotenv
import os
import json
import logging
import psycopg2
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import pytz
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.date import DateTrigger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Загрузка переменных окружения
load_dotenv()
API_TOKEN = os.getenv('API_TOKEN')

# ИСПРАВЛЕНО: Список админов (разделенные запятой в .env)
ADMIN_IDS_STR = os.getenv('ADMIN_IDS', os.getenv('ADMIN_ID', '0'))
ADMIN_IDS = [int(id.strip()) for id in ADMIN_IDS_STR.split(',') if id.strip().isdigit()]

WEB_APP_URL = os.getenv('WEB_APP_URL', 'https://china-together.ru')

# Путь к PDF файлу с тарифами
TARIFFS_PDF_PATH = "/home/chinatogether/xlsx-web/pdf-files/china_together_tariffs.pdf"

# Инициализация бота и хранилища состояний
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# Инициализация планировщика задач (работает в московском времени - Europe/Moscow)
scheduler = AsyncIOScheduler(timezone=pytz.timezone('Europe/Moscow'))

# ==================== FSM СОСТОЯНИЯ ДЛЯ РАССЫЛКИ ====================

class BroadcastStates(StatesGroup):
    """Состояния для FSM рассылки"""
    choosing_type = State()
    waiting_for_test_ids = State()
    waiting_for_message = State()
    waiting_for_buttons = State()
    waiting_for_schedule_time = State()
    confirm_broadcast = State()

# FAQ данные
FAQ_DATA = {
    "💰 Минимальная сумма заказа": (
        "💰 <b>Минимальная сумма заказа</b>\n\n"
        "Минимальная сумма заказа составляет <b>3000 юаней</b> на одного поставщика."
    ),
    "⚖️ Минимальный вес": (
        "⚖️ <b>Минимальный вес к отправке</b>\n\n"
        "Минимальный вес для отправки составляет <b>10 кг/место</b>."
    ),
    "📦 Что такое место": (
        "📦 <b>Что такое «Место»?</b>\n\n"
        "Грузовое место — это одна единица для перевозки, состоящая из одной или нескольких коробок товара.\n\n"
        "💡 <b>Пример:</b> товар 200 единиц приезжает от поставщика в 2-х коробках, мы на складе делаем "
        "деревянный каркас, объединяя обе коробки. После чего получается одно грузовое место."
    ),
    "📦 Виды упаковки": (
        "📦 <b>Виды упаковки и стоимость</b>\n\n"
        "• <b>Мешок + скотч:</b> 3$/место\n"
        "• <b>Картонные уголки + мешок + скотч:</b> 8$/место\n"
        "• <b>Деревянный каркас + мешок + скотч:</b> 15$/место\n"
        "• <b>Палета</b> (деревянный поддон + водозащитная пленка + деревянный каркас + скотч): 30$/куб"
    ),
    "🕐 Сроки доставки": (
        "🕐 <b>Сроки доставки по категориям товаров</b>\n\n"
        "📌 <b>Обычные товары:</b>\n"
        "— Быстрое авто: снижена цена📉\n"
        "— Обычное авто: снижена цена📉\n"
        "<b>Сроки доставки:</b>\n"
        "— быстрое авто: 13–16 дней\n"
        "— обычное авто: 17–25 дней\n\n"
        "📌 <b>Одежда:</b>\n"
        "— Быстрое авто: снижена цена📉\n"
        "— Обычное авто: снижена цена📉\n"
        "<b>Сроки доставки:</b>\n"
        "— быстрое авто: 15–18 дней\n"
        "— обычное авто: 20–25 дней\n\n"
        "📌 <b>Обувь:</b>\n"
        "— Быстрое авто: снижена цена📉\n"
        "<b>Сроки доставки:</b>\n"
        "— быстрое авто: 12–15 дней"
    ),
    "📋 Страховка": (
        "📋 <b>Страхование товара</b>\n\n"
        "Страховка действует при потере, краже или частичной пропаже товара. Карго не возмещает стоимость за помятую упаковку.\n\n"
        "<b>Тарифы страхования:</b>\n"
        "• до 20$/кг — <b>1%</b> от стоимости\n"
        "• 20-30$/кг — <b>2%</b> от стоимости\n"
        "• 30-40$/кг — <b>3%</b> от стоимости\n"
        "• свыше 40$/кг — <b>обсуждается индивидуально</b>\n\n"
        "⚠️ <b>Запрещенные к перевозке товары:</b> порошковые вещества, легковоспламеняющиеся материалы, жидкости, табачные изделия, лекарственные препараты, режущие предметы, продукты питания."
    ),
    "💳 Как происходит оплата": (
        "💳 <b>Процесс оплаты</b>\n\n"
        "1️⃣ <b>Сначала оплачивается:</b>\n"
        "• Стоимость товара\n"
        "• Доставка по Китаю\n"
        "• Комиссия за выкуп\n\n"
        "2️⃣ <b>Доставка до России</b> оплачивается по приезду в Москву, по актуальному курсу."
    ),
    "💸 Способы оплаты": (
        "💸 <b>Доступные способы оплаты</b>\n\n"
        "• 🏦 Карты <b>Т-банк (Тинькофф)/ Альфабанк</b>\n"
        "• 📱 <b>СБП</b> (Система быстрых платежей)\n"
        "• 💰 <b>USDT</b>\n"
        "• 💸 <b>Наличными в Москве</b>"
    ),
    "🏢 Куда приезжает товар": (
        "🏢 <b>Место получения в Москве</b>\n\n"
        "Товар приезжает на <b>Южные ворота</b> в Москве."
    ),
    "🚛 Доставка в регионы": (
        "🚛 <b>Доставка в регионы</b>\n\n"
        "✅ Отправляем по <b>всем городам России</b>\n\n"
        "На Южных воротах в Москве представлены все популярные транспортные компании:\n"
        "• ЖелДорЭкспедиция\n"
        "• СДЭК\n"
        "• ПЭК\n"
        "• Деловые линии\n"
        "• Байкал\n"
        "• Мейджик и другие."
    ),
    "ℹ️ Дополнительная информация": (
        "ℹ️ <b>Дополнительная информация для заказчиков</b>\n\n"
        "<b>🔸 Что входит в нашу комиссию:</b>\n"
        "• Предварительный просчёт доставки до Москвы\n"
        "• Согласование цены и сроков отгрузки\n"
        "• Общение с поставщиком по вопросам товара\n"
        "• Размещение и оплата заказа\n"
        "• Контроль доставки по Китаю\n"
        "• Консолидация груза на нашем складе\n"
        "• Фотоотчёт и упаковка товара\n"
        "• Организация доставки до Москвы\n\n"
        "<b>🔸 Сроки:</b>\n"
        "• Отгрузка поставщиком: 2-7 дней\n"
        "• Доставка по Китаю: 1-5 дней\n"
        "• Согласованная отправка: до 8 вечера по китайскому времени\n"
    ),
    "📋 Правила получения груза": (
        "📋 <b>ПРИНИМАЙТЕ ГРУЗ ПРАВИЛЬНО!</b>\n\n"
        "❗ <b>Важно:</b> После получения груза недостаточно просто привезти товар домой и написать, что что-то не хватает.\n\n"
        "<b>🎥 Обязательно:</b>\n"
        "• Снимите <b>видео вскрытия груза</b>\n"
        "• Покажите внешнее состояние упаковки\n"
        "• Зафиксируйте весь процесс от вскрытия до пересчета товара\n\n"
        "<b>⚠️ При повреждении упаковки:</b>\n"
        "Обязательно зафиксируйте на видео при получении у транспортной компании и при наличии повреждения упаковки и следов вскрытия!\n\n"
        "<b>📅 График работы:</b> ПН-ПТ с 10:00 до 18:00 (МСК)\n\n"
        "Без видео - вопросы о возмещении или компенсации рассматриваться не будут."
    )
}

# Конфигурация базы данных
DB_CONFIG = {
    'dbname': os.getenv('DB_NAME', 'delivery_db'),
    'user': os.getenv('DB_USER'), 
    'password': os.getenv('DB_PASSWORD'),
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': os.getenv('DB_PORT', '5432'),
    'connect_timeout': int(os.getenv('DB_TIMEOUT', '10'))
}

# Подключение к базе данных
def connect_to_db():
    try:
        return psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        logger.error(f"Ошибка подключения к БД: {e}")
        return None

# Функция для сохранения действий пользователя
def save_user_action(telegram_id, action, details=None, username=None, first_name=None, last_name=None):
    conn = connect_to_db()
    if not conn:
        return
    
    try:
        cursor = conn.cursor()
        
        # Обновляем/создаем запись в telegram_users
        if username or first_name or last_name:
            cursor.execute("""
                INSERT INTO delivery_test.telegram_users 
                (telegram_id, username, first_name, last_name, last_activity)
                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                ON CONFLICT (telegram_id) 
                DO UPDATE SET 
                    username = COALESCE(EXCLUDED.username, telegram_users.username),
                    first_name = COALESCE(EXCLUDED.first_name, telegram_users.first_name),
                    last_name = COALESCE(EXCLUDED.last_name, telegram_users.last_name),
                    last_activity = CURRENT_TIMESTAMP
            """, (telegram_id, username, first_name, last_name))
        
        # Сохраняем действие в user_actions
        cursor.execute("""
            INSERT INTO delivery_test.user_actions (telegram_user_id, action, details, created_at)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
        """, (telegram_id, action, json.dumps(details) if details else None))
        
        conn.commit()
        cursor.close()
    except Exception as e:
        logger.error(f"Ошибка сохранения действия: {e}")
    finally:
        conn.close()

# ==================== ФУНКЦИИ ДЛЯ РАССЫЛКИ ====================

def is_admin(user_id: int) -> bool:
    """Проверяет, является ли пользователь администратором"""
    return user_id in ADMIN_IDS

def get_all_users() -> List[int]:
    """Получает список всех уникальных пользователей из БД (только с username)"""
    conn = connect_to_db()
    if not conn:
        return []
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT distinct telegram_user_id 
            FROM delivery_test.user_actions ua
            left join delivery_test.telegram_users tu
            on tu.telegram_id::text=ua.telegram_user_id::text
            WHERE username IS NOT NULL
        """)
        users = [row[0] for row in cursor.fetchall()]
        cursor.close()
        logger.info(f"Получено {len(users)} пользователей из БД (с username)")
        return users
    except Exception as e:
        logger.error(f"Ошибка получения пользователей: {e}")
        return []
    finally:
        conn.close()

def get_username_by_id(user_id: int) -> Optional[str]:
    """Получает username по user_id из БД"""
    conn = connect_to_db()
    if not conn:
        return None
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT username 
            FROM delivery_test.telegram_users
            WHERE telegram_id = %s
        """, (user_id,))
        
        result = cursor.fetchone()
        cursor.close()
        
        if result and result[0]:
            return result[0]
        return None
        
    except Exception as e:
        logger.error(f"Ошибка получения username для {user_id}: {e}")
        return None
    finally:
        conn.close()

def find_user_id_by_username(username: str) -> Optional[int]:
    """Ищет user_id по username в базе данных"""
    clean_username = username.lstrip('@').lower()
    
    conn = connect_to_db()
    if not conn:
        return None
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT telegram_id 
            FROM delivery_test.telegram_users
            WHERE LOWER(username) = %s
            ORDER BY last_activity DESC
            LIMIT 1
        """, (clean_username,))
        
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            return int(result[0])
        return None
        
    except Exception as e:
        logger.error(f"Ошибка поиска по username {username}: {e}")
        return None
    finally:
        conn.close()

def parse_recipients(input_text: str, sender_id: int) -> Tuple[List[int], List[str]]:
    """Парсит строку с ID и username, возвращает список user_id"""
    user_ids = []
    not_found = []
    
    items = input_text.replace(',', ' ').split()
    
    for item in items:
        item = item.strip()
        if not item:
            continue
        
        if item.lower() == 'me':
            user_ids.append(sender_id)
            continue
        
        if item.isdigit():
            user_ids.append(int(item))
            continue
        
        if item.startswith('@') or item.replace('_', '').isalpha():
            user_id = find_user_id_by_username(item)
            if user_id:
                user_ids.append(user_id)
            else:
                not_found.append(item)
            continue
        
        not_found.append(item)
    
    user_ids = list(set(user_ids))
    
    return user_ids, not_found

def save_broadcast_stats(admin_id: int, total: int, success: int, failed: int, 
                        message_text: str, broadcast_type: str = "immediate", 
                        failed_users: Optional[List[Tuple[int, str, Optional[str]]]] = None):
    """Сохраняет статистику рассылки в БД с деталями"""
    conn = connect_to_db()
    if not conn:
        return
    
    try:
        cursor = conn.cursor()
        
        details = {
            "total": total,
            "success": success,
            "failed": failed,
            "type": broadcast_type,
            "message_preview": message_text[:100]
        }
        
        if failed_users:
            details["failed_details"] = [
                {
                    "user_id": user_id,
                    "reason": reason,
                    "username": username
                }
                for user_id, reason, username in failed_users[:50]
            ]
        
        cursor.execute("""
            INSERT INTO delivery_test.user_actions (telegram_user_id, action, details, created_at)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
        """, (admin_id, "broadcast_completed", json.dumps(details)))
        
        conn.commit()
        cursor.close()
    except Exception as e:
        logger.error(f"Ошибка сохранения статистики рассылки: {e}")
    finally:
        conn.close()

async def send_broadcast_to_users(user_ids: List[int], message_text: str, 
                                 buttons: Optional[List[Dict]] = None, 
                                 broadcast_type: str = "immediate") -> Dict:
    """Отправляет рассылку указанным пользователям с детальным логированием"""
    total = len(user_ids)
    success = 0
    failed = 0
    
    failed_users = []
    success_users = []
    
    keyboard = None
    if buttons:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=btn['text'], url=btn['url'])] for btn in buttons
        ])
    
    for user_id in user_ids:
        try:
            await bot.send_message(
                chat_id=user_id,
                text=message_text,
                reply_markup=keyboard,
                parse_mode="HTML"
            )
            success += 1
            success_users.append(user_id)
            logger.info(f"✅ Рассылка отправлена пользователю {user_id}")
            await asyncio.sleep(0.05)
            
        except Exception as e:
            failed += 1
            error_text = str(e)
            
            if "bot was blocked" in error_text.lower():
                reason = "Заблокировал бота"
            elif "user is deactivated" in error_text.lower():
                reason = "Аккаунт удален"
            elif "chat not found" in error_text.lower():
                reason = "Чат не найден"
            elif "forbidden" in error_text.lower():
                reason = "Доступ запрещен"
            elif "user_id" in error_text.lower():
                reason = "Неверный ID"
            else:
                reason = f"Ошибка: {error_text[:50]}"
            
            username = get_username_by_id(user_id)
            
            failed_users.append((user_id, reason, username))
            logger.error(f"❌ Ошибка отправки пользователю {user_id} (@{username}): {reason}")
    
    # Используем первого админа для сохранения статистики
    admin_id = ADMIN_IDS[0] if ADMIN_IDS else 0
    save_broadcast_stats(
        admin_id, 
        total, 
        success, 
        failed, 
        message_text, 
        broadcast_type,
        failed_users=failed_users
    )
    
    return {
        'total': total,
        'success': success,
        'failed': failed,
        'failed_users': failed_users,
        'success_users': success_users
    }

# Функция настройки команд бота
async def set_bot_commands():
    # Команды для обычных пользователей
    user_commands = [
        BotCommand(command="calculate", description="📊 Рассчитать стоимость доставки"),
        BotCommand(command="order", description="🚚 Запросить рассчет у менеджера"),
        BotCommand(command="tariffs", description="📋 Актуальные тарифы"),
        BotCommand(command="help", description="❓ Помощь"),
        BotCommand(command="faq", description="❓ Популярные вопросы"),
        BotCommand(command="restart", description="🔄 Перезапустить бота"),
        BotCommand(command="feedback", description="📝 Отзывы"),
        BotCommand(command="channel", description="📎 Наш канал")
    ]
    
    # Команды для администраторов
    admin_commands = user_commands + [
        BotCommand(command="broadcast", description="📢 Создать рассылку"),
        BotCommand(command="scheduled", description="📅 Запланированные рассылки"),
        BotCommand(command="stats", description="📊 Статистика"),
        BotCommand(command="broadcast_history", description="📚 История рассылок"),
        BotCommand(command="find", description="🔍 Экспорт пользователей CSV"),
        BotCommand(command="export_orders", description="📋 Экспорт заявок XLSX"),
    ]
    
    # Устанавливаем команды для обычных пользователей
    await bot.set_my_commands(user_commands)
    
    # Устанавливаем команды для каждого администратора
    if ADMIN_IDS:
        for admin_id in ADMIN_IDS:
            try:
                await bot.set_my_commands(
                    admin_commands,
                    scope=BotCommandScopeChat(chat_id=admin_id)
                )
                logger.info(f"✅ Команды админа установлены для {admin_id}")
            except Exception as e:
                logger.error(f"❌ Ошибка установки команд для админа {admin_id}: {e}")
    
    logger.info("✅ Команды бота установлены")

# Создание reply-клавиатуры
def get_main_reply_keyboard():
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📊 Рассчитать стоимость доставки"),
                KeyboardButton(text="🚚 Запросить рассчет у менеджера")
            ],
            [
                KeyboardButton(text="📋 Актуальные тарифы"),
                KeyboardButton(text="❓ Помощь")
            ],
            [
                KeyboardButton(text="❓ Популярные вопросы"),
                KeyboardButton(text="🔄 Перезапустить бота")
            ],
            [
                KeyboardButton(text="📝 Отзывы"),
                KeyboardButton(text="📎 Наш канал")
            ]
        ],
        resize_keyboard=True,
        one_time_keyboard=False,
        input_field_placeholder="Выберите действие или используйте меню бота..."
    )
    return keyboard

# Создание inline-клавиатуры для FAQ
def get_faq_inline_keyboard():
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💰 Минимальная сумма заказа", callback_data="faq_min_order")],
        [InlineKeyboardButton(text="⚖️ Минимальный вес", callback_data="faq_min_weight")],
        [InlineKeyboardButton(text="📦 Что такое место", callback_data="faq_place")],
        [InlineKeyboardButton(text="📦 Виды упаковки", callback_data="faq_packaging")],
        [InlineKeyboardButton(text="🕐 Сроки доставки", callback_data="faq_delivery_times")],
        [InlineKeyboardButton(text="📋 Страховка", callback_data="faq_insurance")],
        [InlineKeyboardButton(text="💳 Как происходит оплата", callback_data="faq_payment")],
        [InlineKeyboardButton(text="💸 Способы оплаты", callback_data="faq_payment_methods")],
        [InlineKeyboardButton(text="🏢 Куда приезжает товар", callback_data="faq_delivery_location")],
        [InlineKeyboardButton(text="🚛 Доставка в регионы", callback_data="faq_regions")],
        [InlineKeyboardButton(text="ℹ️ Дополнительная информация", callback_data="faq_additional_info")],
        [InlineKeyboardButton(text="📋 Правила получения груза", callback_data="faq_cargo_rules")]
    ])
    return keyboard

# ИСПРАВЛЕНО: Создание inline-клавиатуры для веб-приложений
def get_webapp_inline_keyboard(user_id, username, action="calculate"):
    if action == "calculate":
        web_app_url = f"{WEB_APP_URL}/calculate?telegram_id={user_id}&username={username}"
        button_text = "📊 Открыть калькулятор"
    else:  # order
        web_app_url = f"{WEB_APP_URL}/order?telegram_id={user_id}&username={username}"
        button_text = "🚚 Открыть форму заказа"
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=button_text, 
            web_app=WebAppInfo(url=web_app_url)
        )]
    ])
    return keyboard

# Функция для отправки PDF с тарифами
async def send_tariffs_pdf(message: types.Message, user_id: int):
    """Отправляет PDF файл с актуальными тарифами"""
    try:
        if not os.path.exists(TARIFFS_PDF_PATH):
            await message.reply(
                "❌ <b>Файл с тарифами не найден</b>\n\n"
                "Обратитесь к менеджеру для получения актуальных тарифов:",
                parse_mode="HTML",
                disable_web_page_preview=True
            )
            save_user_action(user_id, "tariffs_file_not_found")
            return
        
        loading_message = await message.reply(
            "📋 <b>Загружаю актуальные тарифы...</b>\n\n"
            "⏳ Подождите немного, файл готовится к отправке.",
            parse_mode="HTML",
            disable_web_page_preview=True
        )
        
        document = FSInputFile(TARIFFS_PDF_PATH)
        
        await message.reply_document(
            document=document,
            caption=(
                "📋 <b>Актуальные тарифы China Together</b>\n"
            ),
            parse_mode="HTML",
            disable_web_page_preview=True
        )
        
        await loading_message.delete()
        
        save_user_action(user_id, "tariffs_downloaded", {
            "file_path": TARIFFS_PDF_PATH,
            "file_size": os.path.getsize(TARIFFS_PDF_PATH)
        })
        
        logger.info(f"Пользователь {user_id} скачал тарифы")
        
    except Exception as e:
        logger.error(f"Ошибка отправки PDF тарифов: {e}")
        await message.reply(
            "❌ <b>Произошла ошибка при отправке тарифов</b>\n\n"
            "Попробуйте позже или обратитесь к менеджеру:",
            parse_mode="HTML",
            disable_web_page_preview=True
        )
        save_user_action(user_id, "tariffs_send_error", {"error": str(e)})

# Функция для умных ответов ИИ
def get_smart_response(message_text):
    text = message_text.lower()
    
    if any(word in text for word in ['привет', 'здравствуй', 'добро пожаловать', 'hi', 'hello']):
        return ("👋 Привет! Я помогу Вам рассчитать стоимость доставки, оформить заказ и ответить на любые вопросы.\n\n"
                "Используйте кнопки ниже для быстрого доступа к функциям!")
    
    elif any(word in text for word in ['доставка', 'расчет', 'стоимость', 'цена',  'калькулятор']):
        return ("📊 <b>Расчет стоимости доставки:</b>\n\n"
                "Для точного расчета нажмите кнопку '📊 Рассчитать стоимость доставки'. Вам нужно указать:\n"
                "• 📦 Категорию товара (обычные, текстиль, одежда, обувь)\n"
                "• ⚖️ Вес и размеры каждой коробки\n"
                "• 💰 Стоимость товара\n"
                "• 📦 Количество коробок\n\n"
                "У нас 3 варианта упаковки: мешок, картонные уголки, деревянный каркас.\n"
                "2 варианта доставки: быстрая (12-15 дней) и обычная (15-20 дней).\n\n"
                "📋 Для просмотра всех тарифов нажмите кнопку '📋 Актуальные тарифы'")
    
    elif any(word in text for word in ['тариф', 'тарифы', 'прайс', 'цены', 'стоимость']):
        return ("📋 <b>Актуальные тарифы:</b>\n\n"
                "Для получения полного прайс-листа нажмите кнопку '📋 Актуальные тарифы' - вам будет отправлен PDF файл со всеми актуальными ценами.\n")
    
    elif any(word in text for word in ['спасибо', 'благодар', 'thanks', 'отлично', 'хорошо']):
        return ("😊 Пожалуйста! Рады помочь! China Together всегда к вашим услугам. "
                "Если есть еще вопросы - обращайтесь.\n")
    
    else:
        return ("🤔 <b>Я готов помочь!</b>\n\n"
                "🎯 <b>Мои возможности:</b>\n"
                "📊 Расчет стоимости доставки\n"
                "🚚 Оформление заказов и выкуп товаров\n"
                "📋 Предоставление актуальных тарифов\n"
                "❓ Ответы на популярные вопросы\n"
                "❓ Справка и помощь\n\n"
                "Используйте кнопки ниже или меню бота для быстрого доступа к функциям!")

# Функция очистки истории чата
async def clear_chat_history(message: types.Message, user_id: int):
    """Полная очистка истории чата"""
    try:
        current_message_id = message.message_id
        
        deleted_count = 0
        for i in range(1, 501):
            try:
                await bot.delete_message(chat_id=message.chat.id, message_id=current_message_id - i)
                deleted_count += 1
            except Exception as e:
                continue
        
        try:
            await bot.delete_message(chat_id=message.chat.id, message_id=current_message_id)
            deleted_count += 1
        except:
            pass
            
        logger.info(f"Удалено сообщений для пользователя {user_id}: {deleted_count}")
        
    except Exception as e:
        logger.warning(f"Не удалось полностью очистить историю чата для пользователя {user_id}: {e}")

# Функция получения заявок пользователя
def get_user_purchase_requests(telegram_id):
    conn = connect_to_db()
    if not conn:
        return []
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT pr.*, u.telegram_id
            FROM delivery_test.purchase_requests pr
            JOIN delivery_test.telegram_users u ON pr.telegram_user_id = u.id
            WHERE u.telegram_id = %s
            ORDER BY pr.created_at DESC
            LIMIT 10
        """, (str(telegram_id),))
        
        columns = [desc[0] for desc in cursor.description]
        results = []
        for row in cursor.fetchall():
            record = dict(zip(columns, row))
            record['created_at'] = record['created_at'].strftime('%Y-%m-%d %H:%M')
            results.append(record)
        
        cursor.close()
        return results
    except Exception as e:
        logger.error(f"Ошибка получения заявок: {e}")
        return []
    finally:
        conn.close()

# ==================== КОМАНДЫ ПОЛЬЗОВАТЕЛЕЙ ====================

# Команда /start
@dp.message(Command("start"))
async def start(message: types.Message):
    user_id = message.from_user.id
    username = message.from_user.username
    first_name = message.from_user.first_name or ""
    last_name = message.from_user.last_name or ""
    
    save_user_action(
        user_id, 
        "start_command", 
        {"first_name": first_name},
        username=username,
        first_name=first_name,
        last_name=last_name
    )
    
    keyboard = get_main_reply_keyboard()
    
    await message.reply(
        f"🚀 <b>Добро пожаловать, {first_name}!</b>\n\n"
        "Я помогу Вам рассчитать стоимость доставки из Китая и оформить заказ!\n\n"
        "🎯 <b>Что я делаю для вашего удобства:</b>\n"
        "• 📊 Рассчитываю стоимость доставки\n"
        "• 🚚 Оформляю заявки на поиск поставщика и выкуп товара\n"
        "• 📋 Предоставляю актуальные тарифы\n"
        "• 💬 Быстро и по делу отвечаю на частые вопросы\n\n"
        "💡 <b>Используйте кнопки ниже или меню бота для быстрого доступа к функциям!</b>",
        reply_markup=keyboard,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Команда /calculate
@dp.message(Command("calculate"))
async def calculate_delivery(message: types.Message):
    user_id = message.from_user.id
    username = message.from_user.username or f"user_{user_id}"
    
    save_user_action(user_id, "calculate_command")
    
    keyboard = get_webapp_inline_keyboard(user_id, username, "calculate")
    await message.reply(
        "📊 <b>Калькулятор доставки</b>\n\n"
        "Нажмите кнопку ниже, чтобы открыть калькулятор и самостоятельно рассчитать стоимость доставки вашего груза из Китая.\n",
        reply_markup=keyboard,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Команда /order
@dp.message(Command("order"))
async def order_delivery(message: types.Message):
    user_id = message.from_user.id
    username = message.from_user.username or f"user_{user_id}"
    
    save_user_action(user_id, "order_command")
    
    keyboard = get_webapp_inline_keyboard(user_id, username, "order")
    await message.reply(
        "🚚 <b>Заказ доставки</b>\n\n"
        "Нажмите кнопку ниже, чтобы оформить заявку на выкуп и доставку товаров из Китая.\n\n",
        reply_markup=keyboard,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Команда /tariffs
@dp.message(Command("tariffs"))
async def get_tariffs(message: types.Message):
    user_id = message.from_user.id
    save_user_action(user_id, "tariffs_command")
    await send_tariffs_pdf(message, user_id)

# Команда /help
@dp.message(Command("help"))
async def help_command(message: types.Message):
    user_id = message.from_user.id
    save_user_action(user_id, "help_command")
    
    help_text = (
        "<b>❓ Как пользоваться сервисом:</b>\n\n"
        "<b>📊 Расчет доставки (/calculate):</b>\n"
        "1️⃣ Выберите команду в меню\n"
        "2️⃣ Заполните все поля формы\n"
        "3️⃣ Получите детальный расчет\n\n"
        "<b>🚚 Заказ доставки (/order):</b>\n"
        "1️⃣ Выберите команду в меню\n"
        "2️⃣ Заполните заявку\n"
        "3️⃣ Менеджер свяжется с вами\n\n"
        "<b>📋 Актуальные тарифы (/tariffs):</b>\n"
        "Получите PDF с полным прайс-листом\n\n"
        "<b>❓ Популярные вопросы (/faq):</b>\n"
        "Быстрые ответы на частые вопросы\n\n"
        "<b>🔄 Перезапуск (/restart):</b>\n"
        "Очистить чат и начать сначала\n\n"
        "<b>📎 Наш канал (/channel):</b>\n"
        "Полезная информация и новости\n"
        "💡 <b>Все команды доступны через меню бота!</b>"
    )
    await message.reply(help_text, parse_mode="HTML", disable_web_page_preview=True)

# Команда /faq
@dp.message(Command("faq"))
async def faq_command(message: types.Message):
    user_id = message.from_user.id
    save_user_action(user_id, "faq_command")
    
    keyboard = get_faq_inline_keyboard()
    await message.reply(
        "❓ <b>Популярные вопросы</b>\n\n"
        "Выберите интересующий вас вопрос:",
        reply_markup=keyboard,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Команда /feedback
@dp.message(Command("feedback"))
async def feedback_command(message: types.Message):
    user_id = message.from_user.id
    save_user_action(user_id, "feedback_command")
    await message.reply(
        "📝 <b>Отзывы наших клиентов</b>\n\n"
        "🔗 Перейдите по ссылке, чтобы увидеть отзывы:\n"
        "<a href='https://t.me/feedbacktogetherchina'>https://t.me/feedbacktogetherchina</a>\n", 
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Команда /channel
@dp.message(Command("channel"))
async def channel_command(message: types.Message):
    user_id = message.from_user.id
    save_user_action(user_id, "channel_command")
    await message.reply(
        "📎 <b>Наш официальный канал:</b>\n\n"
        "<a href='https://t.me/Togetherchina'>t.me/Togetherchina</a>", 
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Команда /restart
@dp.message(Command("restart"))
async def restart_bot(message: types.Message):
    user_id = message.from_user.id
    first_name = message.from_user.first_name or ""
    
    save_user_action(user_id, "restart_command")
    
    await clear_chat_history(message, user_id)
    
    keyboard = get_main_reply_keyboard()
    
    await bot.send_message(
        chat_id=message.chat.id,
        text=(
            f"🚀 <b>Добро пожаловать, {first_name}!</b>\n\n"
            "Я помогу Вам рассчитать стоимость доставки из Китая и оформить заказ!\n\n"
            "🎯 <b>Что я делаю для вашего удобства:</b>\n"
            "• 📊 Рассчитываю стоимость доставки\n"
            "• 🚚 Оформляю заявки на поиск поставщика и выкуп товара\n"
            "• 📋 Предоставляю актуальные тарифы\n"
            "• 💬 Быстро и по делу отвечаю на частые вопросы\n\n"
            "💡 <b>Используйте кнопки ниже или меню бота для быстрого доступа к функциям!</b>"
        ),
        reply_markup=keyboard,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Обработка reply-кнопок
@dp.message(F.text.in_([
    "📊 Рассчитать стоимость доставки", 
    "🚚 Запросить рассчет у менеджера", 
    "📋 Актуальные тарифы",
    "❓ Помощь",
    "❓ Популярные вопросы",
    "🔄 Перезапустить бота",
    "📝 Отзывы",
    "📎 Наш канал"
]))
async def handle_reply_buttons(message: types.Message):
    user_id = message.from_user.id
    text = message.text
    
    save_user_action(user_id, "reply_button", {"button": text})
    
    if text == "📊 Рассчитать стоимость доставки":
        await calculate_delivery(message)
    elif text == "🚚 Запросить рассчет у менеджера":
        await order_delivery(message)
    elif text == "📋 Актуальные тарифы":
        await get_tariffs(message)
    elif text == "❓ Помощь":
        await help_command(message)
    elif text == "❓ Популярные вопросы":
        await faq_command(message)
    elif text == "🔄 Перезапустить бота":
        await restart_bot(message)
    elif text == "📝 Отзывы":
        await feedback_command(message)
    elif text == "📎 Наш канал":
        await channel_command(message)

# ==================== КОМАНДЫ АДМИНИСТРАТОРА ====================
# ВАЖНО: Команды администратора должны быть ПЕРЕД обработчиком текста!

# Команда /broadcast
@dp.message(Command("broadcast"))
async def broadcast_command(message: types.Message, state: FSMContext):
    """Создание рассылки"""
    if not is_admin(message.from_user.id):
        await message.reply("⛔️ Эта команда доступна только администратору.")
        return
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📢 Отправить всем", callback_data="broadcast_all")],
        [InlineKeyboardButton(text="🧪 Тестовая рассылка", callback_data="broadcast_test")],
        [InlineKeyboardButton(text="⏰ Отложенная рассылка", callback_data="broadcast_scheduled")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="broadcast_cancel")]
    ])
    
    await message.reply(
        "📢 <b>СОЗДАНИЕ РАССЫЛКИ</b>\n\n"
        "Выберите тип рассылки:",
        reply_markup=keyboard,
        parse_mode="HTML"
    )
    await state.set_state(BroadcastStates.choosing_type)

# Команда /cancel
@dp.message(Command("cancel"), StateFilter("*"))
async def cancel_command(message: types.Message, state: FSMContext):
    """Отмена текущего действия"""
    current_state = await state.get_state()
    if current_state is None:
        await message.answer("❌ Нет активных действий для отмены.")
        return
    
    await state.clear()
    await message.answer(
        "❌ <b>Действие отменено</b>\n\n"
        "Используйте /broadcast для создания новой рассылки",
        parse_mode="HTML"
    )

# Выбор типа рассылки - Всем
@dp.callback_query(F.data == "broadcast_all", BroadcastStates.choosing_type)
async def broadcast_all_callback(callback: CallbackQuery, state: FSMContext):
    """Рассылка всем пользователям"""
    await state.update_data(broadcast_type='all')
    await callback.message.edit_text(
        "📢 <b>РАССЫЛКА ВСЕМ ПОЛЬЗОВАТЕЛЯМ</b>\n\n"
        "📝 Отправьте текст сообщения:\n\n"
        "ℹ️ Можно использовать HTML-форматирование:\n"
        "• <b>жирный</b>\n"
        "• <i>курсив</i>\n"
        "• <code>моноширинный</code>\n\n"
        "Для отмены отправьте /cancel",
        parse_mode="HTML"
    )
    await state.set_state(BroadcastStates.waiting_for_message)
    await callback.answer()

# Выбор типа рассылки - Тестовая
@dp.callback_query(F.data == "broadcast_test", BroadcastStates.choosing_type)
async def broadcast_test_callback(callback: CallbackQuery, state: FSMContext):
    """Тестовая рассылка"""
    await state.update_data(broadcast_type='test')
    await callback.message.edit_text(
        "🧪 <b>ТЕСТОВАЯ РАССЫЛКА</b>\n\n"
        "👥 Введите получателей (через запятую или пробел):\n\n"
        "📝 <b>Можно указать:</b>\n"
        "• ID пользователя: <code>123456789</code>\n"
        "• Username: <code>@username</code> или <code>username</code>\n"
        "• Себе: <code>me</code>\n"
        "• Несколько: <code>123456789, @user1, me</code>\n\n"
        "Для отмены отправьте /cancel",
        parse_mode="HTML"
    )
    await state.set_state(BroadcastStates.waiting_for_test_ids)
    await callback.answer()

# Выбор типа рассылки - Отложенная
@dp.callback_query(F.data == "broadcast_scheduled", BroadcastStates.choosing_type)
async def broadcast_scheduled_callback(callback: CallbackQuery, state: FSMContext):
    """Отложенная рассылка"""
    await state.update_data(broadcast_type='scheduled')
    await callback.message.edit_text(
        "⏰ <b>ОТЛОЖЕННАЯ РАССЫЛКА</b>\n\n"
        "📝 Отправьте текст сообщения:\n\n"
        "ℹ️ Можно использовать HTML-форматирование\n\n"
        "Для отмены отправьте /cancel",
        parse_mode="HTML"
    )
    await state.set_state(BroadcastStates.waiting_for_message)
    await callback.answer()

# Отмена рассылки
@dp.callback_query(F.data == "broadcast_cancel")
async def broadcast_cancel_callback(callback: CallbackQuery, state: FSMContext):
    """Отмена рассылки"""
    await state.clear()
    await callback.message.edit_text("❌ Создание рассылки отменено.")
    await callback.answer()

# Получение ID для тестовой рассылки
@dp.message(BroadcastStates.waiting_for_test_ids)
async def process_test_ids(message: types.Message, state: FSMContext):
    """Обработка ID для тестовой рассылки"""
    if message.text == "/cancel":
        await state.clear()
        await message.reply("❌ Создание рассылки отменено.")
        return
    
    user_ids, not_found = parse_recipients(message.text, message.from_user.id)
    
    if not user_ids and not not_found:
        await message.reply("❌ Не удалось распознать ни одного получателя. Попробуйте еще раз.")
        return
    
    response_text = ""
    
    if user_ids:
        await state.update_data(test_ids=user_ids)
        response_text += f"✅ Добавлено <b>{len(user_ids)}</b> получателей\n\n"
    
    if not_found:
        response_text += f"⚠️ <b>Не найдено ({len(not_found)}):</b>\n"
        for item in not_found[:10]:
            response_text += f"• {item}\n"
        if len(not_found) > 10:
            response_text += f"• ... и еще {len(not_found) - 10}\n"
        response_text += "\n"
    
    if not user_ids:
        response_text += "❌ Не удалось найти ни одного пользователя.\n"
        response_text += "Попробуйте еще раз или используйте ID."
        await message.reply(response_text, parse_mode="HTML")
        return
    
    response_text += f"📝 Теперь отправьте текст сообщения.\n"
    response_text += f"Для отмены отправьте /cancel"
    
    await message.reply(response_text, parse_mode="HTML")
    await state.set_state(BroadcastStates.waiting_for_message)

# Получение текста сообщения
@dp.message(BroadcastStates.waiting_for_message)
async def process_broadcast_message(message: types.Message, state: FSMContext):
    """Обработка текста рассылки"""
    if message.text == "/cancel":
        await state.clear()
        await message.reply("❌ Создание рассылки отменено.")
        return
    
    message_text = message.text or message.caption
    if not message_text:
        await message.reply("❌ Сообщение не может быть пустым. Попробуйте еще раз:")
        return
    
    await state.update_data(message_text=message_text)
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, добавить кнопки", callback_data="add_buttons")],
        [InlineKeyboardButton(text="⏭️ Нет, продолжить без кнопок", callback_data="skip_buttons")]
    ])
    
    await message.reply(
        "📝 <b>Текст сохранен!</b>\n\n"
        "🔘 Хотите добавить кнопки к сообщению?",
        reply_markup=keyboard,
        parse_mode="HTML"
    )
    await state.set_state(BroadcastStates.waiting_for_buttons)

# Добавление кнопок
@dp.callback_query(F.data == "add_buttons", BroadcastStates.waiting_for_buttons)
async def add_buttons_callback(callback: CallbackQuery, state: FSMContext):
    """Добавление кнопок"""
    await callback.message.edit_text(
        "🔘 <b>ДОБАВЛЕНИЕ КНОПОК</b>\n\n"
        "Отправьте кнопки в формате:\n"
        "<code>Текст кнопки 1 | https://example.com\n"
        "Текст кнопки 2 | https://example2.com</code>\n\n"
        "📌 <b>Пример:</b>\n"
        "<code>📊 Калькулятор | https://china-together.ru/calc\n"
        "💬 Связаться | https://t.me/username</code>\n\n"
        "Для отмены отправьте /cancel",
        parse_mode="HTML"
    )
    await callback.answer()

# Пропустить добавление кнопок
@dp.callback_query(F.data == "skip_buttons", BroadcastStates.waiting_for_buttons)
async def skip_buttons_callback(callback: CallbackQuery, state: FSMContext):
    """Пропустить добавление кнопок"""
    await state.update_data(buttons=[])
    data = await state.get_data()
    broadcast_type = data.get('broadcast_type', 'all')
    
    if broadcast_type == 'scheduled':
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⏰ Через 10 минут", callback_data="schedule_10m")],
            [InlineKeyboardButton(text="⏰ Через 30 минут", callback_data="schedule_30m")],
            [InlineKeyboardButton(text="⏰ Через 1 час", callback_data="schedule_1h")],
            [InlineKeyboardButton(text="⏰ Через 3 часа", callback_data="schedule_3h")],
            [InlineKeyboardButton(text="🕐 Завтра в 10:00", callback_data="schedule_tomorrow_10")],
            [InlineKeyboardButton(text="🕐 Завтра в 14:00", callback_data="schedule_tomorrow_14")],
            [InlineKeyboardButton(text="⌨️ Ввести свое время", callback_data="schedule_custom")]
        ])
        
        await callback.message.edit_text(
            "⏰ <b>ВЫБОР ВРЕМЕНИ ОТПРАВКИ</b>\n\n"
            "Когда отправить рассылку?",
            reply_markup=keyboard,
            parse_mode="HTML"
        )
        await state.set_state(BroadcastStates.waiting_for_schedule_time)
    else:
        await show_broadcast_preview(callback.message, state)
    
    await callback.answer()

# Обработка кнопок
@dp.message(BroadcastStates.waiting_for_buttons)
async def process_buttons(message: types.Message, state: FSMContext):
    """Обработка кнопок"""
    if message.text == "/cancel":
        await state.clear()
        await message.reply("❌ Создание рассылки отменено.")
        return
    
    buttons = []
    lines = message.text.strip().split('\n')
    
    for line in lines:
        if '|' not in line:
            continue
        
        parts = line.split('|', 1)
        if len(parts) == 2:
            text = parts[0].strip()
            url = parts[1].strip()
            if text and url:
                buttons.append({'text': text, 'url': url})
    
    if not buttons:
        await message.reply(
            "❌ Не удалось распознать кнопки.\n\n"
            "Используйте формат:\n"
            "<code>Текст | https://example.com</code>",
            parse_mode="HTML"
        )
        return
    
    await state.update_data(buttons=buttons)
    
    data = await state.get_data()
    broadcast_type = data.get('broadcast_type', 'all')
    
    if broadcast_type == 'scheduled':
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⏰ Через 10 минут", callback_data="schedule_10m")],
            [InlineKeyboardButton(text="⏰ Через 30 минут", callback_data="schedule_30m")],
            [InlineKeyboardButton(text="⏰ Через 1 час", callback_data="schedule_1h")],
            [InlineKeyboardButton(text="⏰ Через 3 часа", callback_data="schedule_3h")],
            [InlineKeyboardButton(text="🕐 Завтра в 10:00", callback_data="schedule_tomorrow_10")],
            [InlineKeyboardButton(text="🕐 Завтра в 14:00", callback_data="schedule_tomorrow_14")],
            [InlineKeyboardButton(text="⌨️ Ввести свое время", callback_data="schedule_custom")]
        ])
        
        await message.reply(
            "⏰ <b>ВЫБОР ВРЕМЕНИ ОТПРАВКИ</b>\n\n"
            "Когда отправить рассылку?",
            reply_markup=keyboard,
            parse_mode="HTML"
        )
        await state.set_state(BroadcastStates.waiting_for_schedule_time)
    else:
        await show_broadcast_preview(message, state)

# Обработчики выбора времени
@dp.callback_query(F.data.startswith("schedule_"), BroadcastStates.waiting_for_schedule_time)
async def process_schedule_time(callback: CallbackQuery, state: FSMContext):
    """Обработка выбора времени отправки (ВАЖНО: все времена в московском часовом поясе!)"""
    action = callback.data.replace("schedule_", "")
    
    # Используем московский часовой пояс для всех операций
    moscow_tz = pytz.timezone('Europe/Moscow')
    now = datetime.now(moscow_tz)
    
    if action == "10m":
        scheduled_time = now + timedelta(minutes=10)
    elif action == "30m":
        scheduled_time = now + timedelta(minutes=30)
    elif action == "1h":
        scheduled_time = now + timedelta(hours=1)
    elif action == "3h":
        scheduled_time = now + timedelta(hours=3)
    elif action == "tomorrow_10":
        tomorrow = now + timedelta(days=1)
        scheduled_time = tomorrow.replace(hour=10, minute=0, second=0, microsecond=0)
    elif action == "tomorrow_14":
        tomorrow = now + timedelta(days=1)
        scheduled_time = tomorrow.replace(hour=14, minute=0, second=0, microsecond=0)
    elif action == "custom":
        await callback.message.edit_text(
            "⌨️ <b>ВВОД ВРЕМЕНИ</b>\n\n"
            "Введите дату и время в формате:\n"
            "<code>ДД.ММ.ГГГГ ЧЧ:ММ</code>\n\n"
            "📌 <b>Примеры:</b>\n"
            "• <code>06.12.2024 15:30</code>\n"
            "• <code>10.12.2024 09:00</code>\n\n"
            "⏰ Время по Москве (МСК)\n\n"
            "Для отмены отправьте /cancel",
            parse_mode="HTML"
        )
        await callback.answer()
        return
    else:
        await callback.answer("❌ Неизвестная команда")
        return
    
    await state.update_data(scheduled_time=scheduled_time)
    await show_broadcast_preview(callback.message, state)
    await callback.answer()

# Обработка ручного ввода времени
@dp.message(BroadcastStates.waiting_for_schedule_time)
async def process_custom_time(message: types.Message, state: FSMContext):
    """Обработка ручного ввода времени"""
    if message.text == "/cancel":
        await state.clear()
        await message.reply("❌ Создание рассылки отменено.")
        return
    
    try:
        moscow_tz = pytz.timezone('Europe/Moscow')
        scheduled_time = datetime.strptime(message.text.strip(), "%d.%m.%Y %H:%M")
        scheduled_time = moscow_tz.localize(scheduled_time)
        
        now = datetime.now(moscow_tz)
        if scheduled_time <= now:
            await message.reply(
                "❌ Время должно быть в будущем!\n\n"
                "Попробуйте еще раз:"
            )
            return
        
        await state.update_data(scheduled_time=scheduled_time)
        await show_broadcast_preview(message, state)
        
    except ValueError:
        await message.reply(
            "❌ Неверный формат!\n\n"
            "Используйте: <code>ДД.ММ.ГГГГ ЧЧ:ММ</code>\n"
            "Пример: <code>06.12.2024 15:30</code>",
            parse_mode="HTML"
        )

# Показ превью рассылки
async def show_broadcast_preview(message: types.Message, state: FSMContext):
    """Показывает превью рассылки перед отправкой"""
    data = await state.get_data()
    message_text = data.get('message_text', '')
    buttons = data.get('buttons', [])
    broadcast_type = data.get('broadcast_type', 'all')
    test_ids = data.get('test_ids', [])
    scheduled_time = data.get('scheduled_time')
    
    if broadcast_type == 'test':
        recipient_count = len(test_ids)
    else:
        recipient_count = len(get_all_users())
    
    preview_text = "👀 <b>ПРЕВЬЮ РАССЫЛКИ</b>\n\n"
    
    if broadcast_type == 'all':
        preview_text += f"📢 Тип: <b>Всем пользователям</b>\n"
    elif broadcast_type == 'test':
        preview_text += f"🧪 Тип: <b>Тестовая рассылка</b>\n"
    elif broadcast_type == 'scheduled':
        preview_text += f"⏰ Тип: <b>Отложенная рассылка</b>\n"
    
    preview_text += f"👥 Получателей: <b>{recipient_count}</b>\n"
    
    if scheduled_time:
        preview_text += f"⏰ Отправка: <b>{scheduled_time.strftime('%d.%m.%Y %H:%M')}</b> (МСК)\n"
    
    preview_text += f"\n📝 <b>Сообщение:</b>\n{'-'*30}\n{message_text}\n{'-'*30}\n"
    
    if buttons:
        preview_text += f"\n🔘 Кнопки: {len(buttons)} шт.\n"
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, отправить", callback_data="broadcast_confirm")],
        [InlineKeyboardButton(text="❌ Отменить", callback_data="broadcast_cancel")]
    ])
    
    preview_keyboard = None
    if buttons:
        preview_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=btn['text'], url=btn['url'])] for btn in buttons
        ])
    
    await message.answer(
        f"<b>Вот как будет выглядеть сообщение:</b>\n\n{message_text}",
        reply_markup=preview_keyboard,
        parse_mode="HTML"
    )
    
    await message.answer(
        preview_text,
        reply_markup=keyboard,
        parse_mode="HTML"
    )
    
    await state.set_state(BroadcastStates.confirm_broadcast)

# Подтверждение рассылки
@dp.callback_query(F.data == "broadcast_confirm", BroadcastStates.confirm_broadcast)
async def confirm_broadcast(callback: CallbackQuery, state: FSMContext):
    """Подтверждение и отправка рассылки"""
    data = await state.get_data()
    message_text = data.get('message_text', '')
    buttons = data.get('buttons', [])
    broadcast_type = data.get('broadcast_type', 'all')
    test_ids = data.get('test_ids', [])
    scheduled_time = data.get('scheduled_time')
    
    if broadcast_type == 'test':
        user_ids = test_ids
    else:
        user_ids = get_all_users()
    
    if broadcast_type == 'scheduled':
        job_id = f"broadcast_{datetime.now().timestamp()}"
        
        scheduler.add_job(
            send_broadcast_to_users,
            trigger=DateTrigger(run_date=scheduled_time),
            args=[user_ids, message_text, buttons, "scheduled"],
            id=job_id,
            name=f"Broadcast: {message_text[:30]}..."
        )
        
        await callback.message.edit_text(
            f"✅ <b>Рассылка запланирована!</b>\n\n"
            f"⏰ Отправка: <b>{scheduled_time.strftime('%d.%m.%Y %H:%M')}</b> (МСК)\n"
            f"👥 Получателей: <b>{len(user_ids)}</b>\n\n"
            f"🔍 ID задачи: <code>{job_id}</code>\n\n"
            f"Для просмотра запланированных рассылок: /scheduled",
            parse_mode="HTML"
        )
        save_user_action(ADMIN_IDS[0] if ADMIN_IDS else 0, "broadcast_scheduled", {
            "job_id": job_id,
            "scheduled_time": scheduled_time.isoformat(),
            "recipients": len(user_ids),
            "message_preview": message_text[:100]
        })
        
    else:
        await callback.message.edit_text("⏳ <b>Начинаю рассылку...</b>", parse_mode="HTML")
        
        stats = await send_broadcast_to_users(user_ids, message_text, buttons, broadcast_type)
        
        result_text = (
            "📊 <b>РЕЗУЛЬТАТЫ РАССЫЛКИ</b>\n\n"
            f"✅ Успешно: {stats['success']}\n"
            f"❌ Ошибок: {stats['failed']}\n"
            f"📝 Всего: {stats['total']}\n\n"
            f"📈 Процент доставки: {(stats['success']/stats['total']*100):.1f}%"
        )
        
        await callback.message.answer(result_text, parse_mode="HTML")
        
        if stats['failed'] > 0 and stats['failed_users']:
            failed_text = "❌ <b>ДЕТАЛЬНЫЙ ОТЧЕТ ПО ОШИБКАМ</b>\n\n"
            
            reasons_count = {}
            for user_id, reason, username in stats['failed_users']:
                if reason not in reasons_count:
                    reasons_count[reason] = []
                reasons_count[reason].append((user_id, username))
            
            for reason, users in reasons_count.items():
                failed_text += f"<b>{reason}:</b> {len(users)} чел.\n"
                for user_id, username in users[:5]:
                    username_str = f"@{username}" if username else "—"
                    failed_text += f"  • {user_id} ({username_str})\n"
                
                if len(users) > 5:
                    failed_text += f"  • ... и еще {len(users) - 5}\n"
                failed_text += "\n"
            
            if len(failed_text) > 4000:
                filename = f"/tmp/broadcast_failed_{callback.from_user.id}.txt"
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("ОТЧЕТ ПО НЕУДАЧНЫМ ОТПРАВКАМ\n")
                    f.write("="*60 + "\n\n")
                    for user_id, reason, username in stats['failed_users']:
                        username_str = f"@{username}" if username else "—"
                        f.write(f"{user_id}\t{username_str}\t{reason}\n")
                
                document = FSInputFile(filename)
                await callback.message.answer_document(
                    document=document,
                    caption="📄 Полный отчет по неудачным отправкам"
                )
                
                os.remove(filename)
            else:
                await callback.message.answer(failed_text, parse_mode="HTML")
    
    await state.clear()
    await callback.answer()

# Команда /stats
@dp.message(Command("stats"))
async def stats_command(message: types.Message):
    """Показывает статистику бота"""
    if not is_admin(message.from_user.id):
        await message.reply("⛔️ Эта команда доступна только администратору.")
        return
    
    conn = connect_to_db()
    if not conn:
        await message.reply("❌ Ошибка подключения к БД")
        return
    
    try:
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(DISTINCT telegram_id) 
            FROM delivery_test.telegram_users
        """)
        user_count = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT action, COUNT(*) as count
            FROM delivery_test.user_actions
            GROUP BY action
            ORDER BY count DESC
            LIMIT 10
        """)
        top_actions = cursor.fetchall()
        
        scheduled_count = len(scheduler.get_jobs())
        
        cursor.close()
        
        stats_text = (
            "📊 <b>СТАТИСТИКА БОТА</b>\n\n"
            f"👥 Всего пользователей: <b>{user_count}</b>\n"
            f"📅 Запланировано рассылок: <b>{scheduled_count}</b>\n\n"
            "🔥 <b>Топ действий:</b>\n"
        )
        
        for action, count in top_actions:
            stats_text += f"• {action}: {count}\n"
        
        await message.reply(stats_text, parse_mode="HTML")
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики: {e}")
        await message.reply("❌ Ошибка получения статистики")
    finally:
        conn.close()

# Команда /scheduled
@dp.message(Command("scheduled"))
async def scheduled_command(message: types.Message):
    """Показывает запланированные рассылки"""
    if not is_admin(message.from_user.id):
        await message.reply("⛔️ Эта команда доступна только администратору.")
        return
    
    jobs = scheduler.get_jobs()
    
    if not jobs:
        await message.reply("📭 Нет запланированных рассылок")
        return
    
    text = "📅 <b>ЗАПЛАНИРОВАННЫЕ РАССЫЛКИ</b>\n\n"
    
    keyboard_buttons = []
    
    for job in jobs:
        job_time = job.next_run_time.strftime('%d.%m.%Y %H:%M')
        text += f"⏰ <b>{job_time}</b> (МСК)\n"
        text += f"📝 {job.name}\n"
        text += f"🔍 ID: <code>{job.id}</code>\n\n"
        
        keyboard_buttons.append([
            InlineKeyboardButton(
                text=f"❌ Отменить {job_time}",
                callback_data=f"cancel_job_{job.id}"
            )
        ])
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)
    
    await message.reply(text, reply_markup=keyboard, parse_mode="HTML")

@dp.callback_query(F.data.startswith("cancel_job_"))
async def cancel_scheduled_job(callback: CallbackQuery):
    """Отмена запланированной рассылки"""
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔️ Только для администратора", show_alert=True)
        return
    
    job_id = callback.data.replace("cancel_job_", "")
    
    try:
        scheduler.remove_job(job_id)
        save_user_action(ADMIN_IDS[0] if ADMIN_IDS else 0, "broadcast_cancelled", {"job_id": job_id})
        
        await callback.message.edit_text(
            f"✅ <b>Рассылка отменена</b>\n\n"
            f"🔍 ID: <code>{job_id}</code>",
            parse_mode="HTML"
        )
        await callback.answer("✅ Рассылка отменена")
    except Exception as e:
        logger.error(f"Ошибка отмены рассылки: {e}")
        await callback.answer("❌ Ошибка отмены", show_alert=True)

# Команда /broadcast_history
@dp.message(Command("broadcast_history"))
async def broadcast_history_command(message: types.Message):
    """Показывает историю последних рассылок"""
    if not is_admin(message.from_user.id):
        await message.reply("⛔️ Эта команда доступна только администратору.")
        return
    
    conn = connect_to_db()
    if not conn:
        await message.reply("❌ Ошибка подключения к БД")
        return
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT details, created_at
            FROM delivery_test.user_actions
            WHERE action = 'broadcast_completed'
            ORDER BY created_at DESC
            LIMIT 10
        """)
        
        results = cursor.fetchall()
        cursor.close()
        
        if not results:
            await message.reply("📭 История рассылок пуста")
            return
        
        text = "📚 <b>ИСТОРИЯ РАССЫЛОК</b>\n\n"
        
        for details, created_at in results:
            details_dict = json.loads(details) if isinstance(details, str) else details
            
            date_str = created_at.strftime('%d.%m %H:%M')
            total = details_dict.get('total', 0)
            success = details_dict.get('success', 0)
            failed = details_dict.get('failed', 0)
            preview = details_dict.get('message_preview', '')[:30]
            broadcast_type = details_dict.get('type', 'immediate')
            
            type_emoji = {
                'immediate': '📢',
                'test': '🧪',
                'scheduled': '⏰'
            }.get(broadcast_type, '📢')
            
            text += f"{type_emoji} <b>{date_str}</b>\n"
            text += f"📝 {preview}...\n"
            text += f"✅ {success}/{total}"
            if failed > 0:
                text += f" ❌ {failed}"
            text += "\n\n"
        
        await message.reply(text, parse_mode="HTML")
        
    except Exception as e:
        logger.error(f"Ошибка получения истории: {e}")
        await message.reply("❌ Ошибка получения истории")
    finally:
        conn.close()

# Команда /find
@dp.message(Command("find"))
async def find_user_command(message: types.Message):
    """Поиск пользователей по username с экспортом в CSV"""
    if not is_admin(message.from_user.id):
        await message.reply("⛔️ Эта команда доступна только администратору.")
        return
    
    conn = connect_to_db()
    if not conn:
        await message.reply("❌ Ошибка подключения к БД")
        return
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT telegram_id, username, first_name, last_name, last_activity, created_at
            FROM delivery_test.telegram_users
            ORDER BY last_activity DESC
        """)
        
        results = cursor.fetchall()
        cursor.close()
        
        if not results:
            await message.reply("📭 Пользователи не найдены")
            return
        
        # Статистика
        total_users = len(results)
        users_with_username = sum(1 for r in results if r[1])
        
        # Создаем CSV файл
        import csv
        from datetime import datetime as dt
        import tempfile
        
        # Используем временную директорию, которая точно существует
        try:
            # Пробуем создать файл в /tmp
            temp_dir = '/tmp'
            if not os.path.exists(temp_dir):
                # Если /tmp нет, используем текущую директорию
                temp_dir = os.getcwd()
            
            filename = os.path.join(temp_dir, f"users_{message.from_user.id}_{int(dt.now().timestamp())}.csv")
            
            # Создаем CSV
            with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['telegram_id', 'username', 'first_name', 'last_name', 'last_activity', 'created_at'])
                
                for row in results:
                    telegram_id, username, first_name, last_name, last_activity, created_at = row
                    writer.writerow([
                        telegram_id,
                        username or '',
                        first_name or '',
                        last_name or '',
                        last_activity.strftime('%Y-%m-%d %H:%M:%S') if last_activity else '',
                        created_at.strftime('%Y-%m-%d %H:%M:%S') if created_at else ''
                    ])
            
            logger.info(f"CSV файл создан: {filename}")
            
        except Exception as e:
            logger.error(f"Ошибка создания CSV файла: {e}")
            await message.reply(f"❌ Ошибка создания CSV файла: {e}")
            return
        
        # Формируем текст для отображения (первые 30)
        text = "👥 <b>БАЗА ПОЛЬЗОВАТЕЛЕЙ</b>\n\n"
        text += f"📊 <b>Статистика:</b>\n"
        text += f"• Всего пользователей: <b>{total_users}</b>\n"
        text += f"• С username: <b>{users_with_username}</b>\n"
        text += f"• Без username: <b>{total_users - users_with_username}</b>\n\n"
        text += f"📄 <b>Первые 30 пользователей:</b>\n\n"
        
        for telegram_id, username, first_name, last_name, last_activity, created_at in results[:30]:
            name_str = first_name or ""
            if last_name:
                name_str += f" {last_name}"
            name_str = name_str.strip() or "—"
            
            username_str = f"@{username}" if username else "—"
            
            text += f"• <code>{telegram_id}</code>\n"
            text += f"  🏷 {username_str}\n"
            text += f"  👤 {name_str}\n"
            if last_activity:
                text += f"  📅 {last_activity.strftime('%d.%m.%Y %H:%M')}\n"
            text += "\n"
        
        if total_users > 30:
            text += f"... и еще {total_users - 30} пользователей\n\n"
        
        text += f"📥 <b>Полная база в CSV файле ниже</b>"
        
        # Отправляем текст
        await message.reply(text, parse_mode="HTML")
        
        # Проверяем что файл существует перед отправкой
        if not os.path.exists(filename):
            await message.reply("❌ Ошибка: CSV файл не найден")
            logger.error(f"CSV файл не найден: {filename}")
            return
        
        # Отправляем CSV файл
        try:
            document = FSInputFile(filename)
            await message.reply_document(
                document=document,
                caption=(
                    f"📊 <b>База пользователей China Together Bot</b>\n\n"
                    f"👥 Всего: {total_users} чел.\n"
                    f"🏷 С username: {users_with_username} чел.\n"
                    f"📅 Экспорт: {dt.now().strftime('%d.%m.%Y %H:%M')}"
                ),
                parse_mode="HTML"
            )
            logger.info(f"CSV файл отправлен админу {message.from_user.id}")
        except Exception as e:
            logger.error(f"Ошибка отправки CSV: {e}")
            await message.reply(f"❌ Ошибка отправки файла: {e}")
            return
        finally:
            # Удаляем временный файл
            try:
                if os.path.exists(filename):
                    os.remove(filename)
                    logger.info(f"CSV файл удален: {filename}")
            except Exception as e:
                logger.error(f"Ошибка удаления временного файла: {e}")
        
        logger.info(f"Админ {message.from_user.id} экспортировал базу пользователей ({total_users} чел.)")
        
    except Exception as e:
        logger.error(f"Ошибка поиска пользователей: {e}")
        await message.reply(f"❌ Ошибка выполнения запроса: {e}")
    finally:
        conn.close()

# Команда /export_orders
@dp.message(Command("export_orders"))
async def export_orders_command(message: types.Message):
    """Экспорт заявок на покупку в XLSX"""
    if not is_admin(message.from_user.id):
        await message.reply("⛔️ Эта команда доступна только администратору.")
        return
    
    # Отправляем сообщение о начале экспорта
    status_msg = await message.reply("⏳ <b>Формирую файл с заявками...</b>", parse_mode="HTML")
    
    conn = connect_to_db()
    if not conn:
        await status_msg.edit_text("❌ Ошибка подключения к БД")
        return
    
    try:
        cursor = conn.cursor()
        
        # Получаем заявки с информацией о клиентах
        query = """
            SELECT 
                pr.id,
                tu.telegram_id as client_telegram_id,
                pr.telegram_contact,
                pr.supplier_link,
                pr.order_amount,
                pr.promo_code,
                pr.additional_notes,
                pr.terms_accepted,
                pr.status,
                pr.manager_notes,
                pr.manager_email,
                pr.created_at,
                pr.updated_at,
                pr.google_form_submitted,
                pr.google_form_submission_time,
                pr.calculation_id
            FROM delivery_test.purchase_requests pr
            LEFT JOIN delivery_test.telegram_users tu ON pr.telegram_user_id = tu.id
            ORDER BY pr.created_at DESC
        """
        
        logger.info(f"Выполнение SQL запроса для экспорта заявок")
        cursor.execute(query)
        
        results = cursor.fetchall()
        cursor.close()
        
        logger.info(f"Получено {len(results)} заявок")
        if results:
            logger.info(f"Количество столбцов в результате: {len(results[0])}")
        
        if not results:
            await status_msg.edit_text("📭 Заявки не найдены в базе данных")
            return
        
        # Создаем XLSX файл
        from datetime import datetime as dt
        
        # Определяем директорию для временного файла
        try:
            temp_dir = '/tmp'
            if not os.path.exists(temp_dir):
                temp_dir = os.getcwd()
            
            filename = os.path.join(temp_dir, f"purchase_requests_{message.from_user.id}_{int(dt.now().timestamp())}.xlsx")
            
            # Создаем workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Заявки"
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Заголовки столбцов
            headers = [
                'ID', 'ID польз. (БД)', 'Telegram ID', 'Username', 'Имя', 'Фамилия',
                'Email', 'Telegram контакт', 'Ссылка на поставщика', 'Сумма заказа',
                'Промокод', 'Доп. заметки', 'Условия приняты', 'Статус',
                'Заметки менеджера', 'Email менеджера', 'Создана', 'Обновлена',
                'Google форма', 'Время отправки формы', 'ID расчета'
            ]
            
            # Записываем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Записываем данные
            for row_num, row_data in enumerate(results, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    # Форматируем datetime
                    if isinstance(value, datetime):
                        cell.value = value.strftime('%d.%m.%Y %H:%M')
                    elif isinstance(value, bool):
                        cell.value = "Да" if value else "Нет"
                    elif value is None:
                        cell.value = ""
                    else:
                        cell.value = str(value)
                    
                    cell.border = border
                    
                    # Выравнивание
                    if col_num in [1, 2, 3, 13, 19, 21]:  # ID колонки и boolean
                        cell.alignment = Alignment(horizontal="center")
                    elif col_num == 10:  # Сумма
                        cell.alignment = Alignment(horizontal="right")
            
            # Автоматическая ширина столбцов
            column_widths = {
                1: 8,   # ID
                2: 12,  # ID польз. БД
                3: 15,  # Telegram ID
                4: 20,  # Username
                5: 15,  # Имя
                6: 15,  # Фамилия
                7: 30,  # Email
                8: 20,  # Telegram контакт
                9: 40,  # Ссылка
                10: 15, # Сумма
                11: 15, # Промокод
                12: 40, # Доп. заметки
                13: 12, # Условия
                14: 15, # Статус
                15: 40, # Заметки менеджера
                16: 30, # Email менеджера
                17: 18, # Создана
                18: 18, # Обновлена
                19: 12, # Google форма
                20: 18, # Время отправки
                21: 12  # ID расчета
            }
            
            for col_num, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            # Закрепляем первую строку
            ws.freeze_panes = 'A2'
            
            # Сохраняем файл
            wb.save(filename)
            logger.info(f"XLSX файл создан: {filename}")
            
        except Exception as e:
            logger.error(f"Ошибка создания XLSX файла: {e}")
            await status_msg.edit_text(f"❌ Ошибка создания XLSX файла: {e}")
            return
        
        # Статистика
        total_requests = len(results)
        
        # Подсчет по статусам (индекс 13 в SQL запросе)
        status_counts = {}
        google_forms_sent = 0
        
        try:
            for row in results:
                # Статус на позиции 13
                if len(row) > 13 and row[13]:
                    status = row[13]
                    status_counts[status] = status_counts.get(status, 0) + 1
                
                # Google форма на позиции 18
                if len(row) > 18 and row[18]:
                    google_forms_sent += 1
                    
        except Exception as e:
            logger.error(f"Ошибка подсчета статистики: {e}, длина row: {len(row) if 'row' in locals() else 'unknown'}")
        
        # Формируем текст со статистикой
        text = "📋 <b>ЭКСПОРТ ЗАЯВОК НА ПОКУПКУ</b>\n\n"
        text += f"📊 <b>Статистика:</b>\n"
        text += f"• Всего заявок: <b>{total_requests}</b>\n"
        text += f"• Google форм отправлено: <b>{google_forms_sent}</b>\n\n"
        
        if status_counts:
            text += f"📈 <b>По статусам:</b>\n"
            for status, count in sorted(status_counts.items(), key=lambda x: x[1], reverse=True):
                status_emoji = {
                    'new': '🆕',
                    'in_progress': '⏳',
                    'completed': '✅',
                    'cancelled': '❌',
                    'rejected': '🚫'
                }.get(status, '📋')
                text += f"  {status_emoji} {status}: <b>{count}</b>\n"
        
        text += f"\n📥 <b>Полная выгрузка в XLSX файле ниже</b>"
        
        # Отправляем статистику
        await status_msg.edit_text(text, parse_mode="HTML")
        
        # Проверяем что файл существует
        if not os.path.exists(filename):
            await message.reply("❌ Ошибка: XLSX файл не найден")
            logger.error(f"XLSX файл не найден: {filename}")
            return
        
        # Отправляем XLSX файл
        try:
            document = FSInputFile(filename)
            await message.reply_document(
                document=document,
                caption=(
                    f"📋 <b>Заявки на покупку China Together Bot</b>\n\n"
                    f"📊 Всего заявок: {total_requests}\n"
                    f"📝 Google форм: {google_forms_sent}\n"
                    f"📅 Экспорт: {dt.now().strftime('%d.%m.%Y %H:%M')}\n\n"
                    f"📄 Формат: XLSX (Excel)"
                ),
                parse_mode="HTML"
            )
            logger.info(f"XLSX файл отправлен админу {message.from_user.id}")
        except Exception as e:
            logger.error(f"Ошибка отправки XLSX: {e}")
            await message.reply(f"❌ Ошибка отправки файла: {e}")
            return
        finally:
            # Удаляем временный файл
            try:
                if os.path.exists(filename):
                    os.remove(filename)
                    logger.info(f"XLSX файл удален: {filename}")
            except Exception as e:
                logger.error(f"Ошибка удаления временного файла: {e}")
        
        logger.info(f"Админ {message.from_user.id} экспортировал заявки ({total_requests} шт.)")
        
    except Exception as e:
        logger.error(f"Ошибка экспорта заявок: {e}")
        await status_msg.edit_text(f"❌ Ошибка экспорта заявок: {e}")
    finally:
        conn.close()

# Удаление webhook
async def delete_webhook():
    await bot.delete_webhook()

# ==================== ОБРАБОТЧИКИ ТЕКСТА И FAQ (ДОЛЖНЫ БЫТЬ В КОНЦЕ!) ====================

# Обработка текстовых сообщений (ВАЖНО: должен быть последним!)
@dp.message(F.text)
async def handle_text_message(message: types.Message):
    user_id = message.from_user.id
    text = message.text
    
    # ВАЖНО: Игнорируем команды (начинающиеся с /)
    if text.startswith('/'):
        return
    
    button_texts = [
        "📊 Рассчитать стоимость доставки", "🚚 Запросить рассчет у менеджера", "📋 Актуальные тарифы",
        "❓ Помощь", "❓ Популярные вопросы", "🔄 Перезапустить бота", 
        "📝 Отзывы", "📎 Наш канал"
    ]
    
    if text in button_texts:
        return
    
    save_user_action(user_id, "text_message", {"text": text})
    
    smart_response = get_smart_response(text)
    keyboard = get_main_reply_keyboard()
    
    await message.reply(
        smart_response,
        reply_markup=keyboard,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

# Обработка callback-запросов FAQ
@dp.callback_query(F.data.startswith("faq_"))
async def handle_faq_callback(callback: CallbackQuery):
    user_id = callback.from_user.id
    
    faq_key_map = {
        "faq_min_order": "💰 Минимальная сумма заказа",
        "faq_min_weight": "⚖️ Минимальный вес",
        "faq_place": "📦 Что такое место",
        "faq_packaging": "📦 Виды упаковки",
        "faq_delivery_times": "🕐 Сроки доставки",
        "faq_insurance": "📋 Страховка",
        "faq_payment": "💳 Как происходит оплата",
        "faq_payment_methods": "💸 Способы оплаты",
        "faq_delivery_location": "🏢 Куда приезжает товар",
        "faq_regions": "🚛 Доставка в регионы",
        "faq_additional_info": "ℹ️ Дополнительная информация",
        "faq_cargo_rules": "📋 Правила получения груза"
    }
    
    faq_key = faq_key_map.get(callback.data)
    if faq_key and faq_key in FAQ_DATA:
        back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 К вопросам", callback_data="back_to_faq")]
        ])
        await callback.message.edit_text(
            FAQ_DATA[faq_key],
            reply_markup=back_keyboard,
            parse_mode="HTML",
            disable_web_page_preview=True
        )
        save_user_action(user_id, "faq_viewed", {"question": faq_key})
    
    await callback.answer()

@dp.callback_query(F.data == "back_to_faq")
async def back_to_faq_callback(callback: CallbackQuery):
    keyboard = get_faq_inline_keyboard()
    await callback.message.edit_text(
        "❓ <b>Популярные вопросы</b>\n\n"
        "Выберите интересующий вас вопрос:",
        reply_markup=keyboard,
        parse_mode="HTML",
        disable_web_page_preview=True
    )
    await callback.answer()

# Запуск бота
async def main():
    await delete_webhook()
    
    scheduler.start()
    logger.info("⏰ Планировщик задач запущен!")
    
    await set_bot_commands()
    
    logger.info("🤖 China Together Bot запущен!")
    logger.info(f"📢 Функция рассылки активирована!")
    logger.info(f"👤 Администраторы: {ADMIN_IDS}")
    logger.info(f"🌐 Web App URL: {WEB_APP_URL}")
    logger.info(f"📋 Tariffs PDF Path: {TARIFFS_PDF_PATH}")
    
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())